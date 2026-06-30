"""
英语真题词频统计 WebApp — Flask 后端
"""

import sqlite3
import json
import os
import io
import webbrowser
import threading
from flask import Flask, jsonify, request, send_file, render_template

app = Flask(__name__)

# 数据库路径（支持环境变量覆盖，方便 PyInstaller 打包后指定路径）
DB_PATH = os.environ.get(
    'EXAM_WORDS_DB_PATH',
    os.path.join(os.path.dirname(__file__), '..', 'data', 'exam_words.db')
)


def get_db():
    """获取数据库连接"""
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    return conn


# ============================================================
# API: 筛选器元数据
# ============================================================
@app.route('/api/meta/filters')
def get_filters():
    """返回筛选器的可选值范围"""
    conn = get_db()
    try:
        # 年份范围
        years = [row[0] for row in
                 conn.execute("SELECT DISTINCT year FROM occurrences ORDER BY year")]

        # 题型列表
        exam_types = [row[0] for row in
                      conn.execute("SELECT DISTINCT exam_type FROM occurrences ORDER BY exam_type")]

        # 词频范围
        freq_range = conn.execute(
            "SELECT MIN(total_freq), MAX(total_freq) FROM words"
        ).fetchone()

        # 年份数范围
        year_count_range = conn.execute(
            "SELECT MIN(year_count), MAX(year_count) FROM words"
        ).fetchone()

        return jsonify({
            'years': years,
            'exam_types': exam_types,
            'freq_range': {'min': freq_range[0], 'max': freq_range[1]},
            'year_count_range': {'min': year_count_range[0], 'max': year_count_range[1]},
            'total_words': conn.execute("SELECT COUNT(*) FROM words").fetchone()[0]
        })
    finally:
        conn.close()


# ============================================================
# API: 词频列表（支持全部筛选参数组合）
# ============================================================
@app.route('/api/words')
def get_words():
    """获取词频列表，支持多条件组合筛选"""
    conn = get_db()
    try:
        # --- 解析筛选参数 ---
        freq_min = request.args.get('freq_min', type=int)
        freq_max = request.args.get('freq_max', type=int)
        years = request.args.get('years')          # 逗号分隔
        year_count_min = request.args.get('year_count_min', type=int)
        year_count_max = request.args.get('year_count_max', type=int)
        first_year = request.args.get('first_year', type=int)
        exam_types = request.args.get('exam_types')  # 逗号分隔
        search = request.args.get('search', type=str)
        page = request.args.get('page', 1, type=int)
        page_size = request.args.get('page_size', 50, type=int)
        sort_by = request.args.get('sort_by', 'freq')
        sort_order = request.args.get('sort_order', 'desc')

        # --- 构建查询 ---
        conditions = []
        params = []

        # 词频范围
        if freq_min is not None:
            conditions.append("w.total_freq >= ?")
            params.append(freq_min)
        if freq_max is not None:
            conditions.append("w.total_freq <= ?")
            params.append(freq_max)

        # 年份数范围
        if year_count_min is not None:
            conditions.append("w.year_count >= ?")
            params.append(year_count_min)
        if year_count_max is not None:
            conditions.append("w.year_count <= ?")
            params.append(year_count_max)

        # 首次出现年份
        if first_year is not None:
            conditions.append("w.first_year = ?")
            params.append(first_year)

        # 单词搜索
        if search and search.strip():
            conditions.append("w.word LIKE ?")
            params.append(f"%{search.strip().lower()}%")

        # 年份筛选（需要在 occurrences 中出现过）
        # 用子查询实现
        if years and years.strip():
            year_list = [y.strip() for y in years.split(',') if y.strip()]
            if year_list:
                placeholders = ','.join(['?'] * len(year_list))
                conditions.append(
                    f"w.id IN (SELECT DISTINCT word_id FROM occurrences WHERE year IN ({placeholders}))"
                )
                params.extend(year_list)

        # 题型筛选
        if exam_types and exam_types.strip():
            type_list = [t.strip() for t in exam_types.split(',') if t.strip()]
            if type_list:
                placeholders = ','.join(['?'] * len(type_list))
                conditions.append(
                    f"w.id IN (SELECT DISTINCT word_id FROM occurrences WHERE exam_type IN ({placeholders}))"
                )
                params.extend(type_list)

        # --- 排序 ---
        sort_map = {
            'freq': 'w.total_freq',
            'year_count': 'w.year_count',
            'first_year': 'w.first_year',
            'word': 'w.word'
        }
        sort_col = sort_map.get(sort_by, 'w.total_freq')
        order = 'DESC' if sort_order == 'desc' else 'ASC'

        where_clause = ' AND '.join(conditions) if conditions else '1=1'

        # --- 查询总数和总词频 ---
        count_sql = f"SELECT COUNT(*), COALESCE(SUM(total_freq), 0) FROM words w WHERE {where_clause}"
        count_row = conn.execute(count_sql, params).fetchone()
        total = count_row[0]
        total_freq_sum = count_row[1]

        # --- 分页查询 ---
        offset = (page - 1) * page_size
        data_sql = f"""
            SELECT w.word, w.total_freq, w.year_count, w.first_year, w.last_year,
                   w.freq_by_type
            FROM words w
            WHERE {where_clause}
            ORDER BY {sort_col} {order}
            LIMIT ? OFFSET ?
        """
        rows = conn.execute(data_sql, params + [page_size, offset]).fetchall()

        # --- 组装结果 ---
        words = []
        for row in rows:
            words.append({
                'word': row['word'],
                'total_freq': row['total_freq'],
                'year_count': row['year_count'],
                'first_year': row['first_year'],
                'last_year': row['last_year'],
                'freq_by_type': json.loads(row['freq_by_type']) if row['freq_by_type'] else {}
            })

        return jsonify({
            'total': total,
            'total_freq_sum': total_freq_sum,
            'page': page,
            'page_size': page_size,
            'total_pages': max(1, (total + page_size - 1) // page_size),
            'words': words
        })
    finally:
        conn.close()


# ============================================================
# API: 单个单词统计详情
# ============================================================
@app.route('/api/words/<word>')
def get_word_detail(word):
    """获取单个单词的统计详情"""
    conn = get_db()
    try:
        row = conn.execute(
            "SELECT * FROM words WHERE word = ?", (word.lower(),)
        ).fetchone()

        if not row:
            return jsonify({'error': '单词不存在'}), 404

        # 分年词频
        freq_by_year = json.loads(row['freq_by_year']) if row['freq_by_year'] else {}

        # 分题型词频
        freq_by_type = json.loads(row['freq_by_type']) if row['freq_by_type'] else {}

        # 出现次数（用于分页）
        occ_count = conn.execute(
            "SELECT COUNT(*) FROM occurrences WHERE word_id = ?", (row['id'],)
        ).fetchone()[0]

        return jsonify({
            'word': row['word'],
            'total_freq': row['total_freq'],
            'year_count': row['year_count'],
            'first_year': row['first_year'],
            'last_year': row['last_year'],
            'freq_by_year': freq_by_year,
            'freq_by_type': freq_by_type,
            'occurrence_count': occ_count
        })
    finally:
        conn.close()


# ============================================================
# API: 单词出现位置（分页，每页10条）
# ============================================================
@app.route('/api/words/<word>/occurrences')
def get_word_occurrences(word):
    """获取单词的所有出现位置，分页返回（每页10条）"""
    conn = get_db()
    try:
        # 查找单词
        word_row = conn.execute(
            "SELECT id FROM words WHERE word = ?", (word.lower(),)
        ).fetchone()

        if not word_row:
            return jsonify({'error': '单词不存在'}), 404

        word_id = word_row['id']

        # 获取总出现次数
        total = conn.execute(
            "SELECT COUNT(*) FROM occurrences WHERE word_id = ?", (word_id,)
        ).fetchone()[0]

        # 解析分页参数
        page = request.args.get('page', 1, type=int)
        page_size = request.args.get('page_size', 10, type=int)
        offset = (page - 1) * page_size

        # 查询
        rows = conn.execute(
            """SELECT * FROM occurrences
               WHERE word_id = ?
               ORDER BY year, exam_type, id
               LIMIT ? OFFSET ?""",
            (word_id, page_size, offset)
        ).fetchall()

        cards = []
        for row in rows:
            # 对于完形填空选项，需要额外获取题目句
            # 查找同一题组中相邻的正文句子（题目句）
            card = {
                'year': row['year'],
                'exam_type': row['exam_type'],
                'section': row['section'],
                'sentence': row['sentence'],
                'offset': row['word_offset'],
                'length': row['word_length'],
                'surface': row['surface']
            }
            cards.append(card)

        return jsonify({
            'word': word.lower(),
            'total': total,
            'page': page,
            'page_size': page_size,
            'total_pages': max(1, (total + page_size - 1) // page_size),
            'cards': cards
        })
    finally:
        conn.close()


# ============================================================
# API: Excel 导出
# ============================================================
@app.route('/api/export')
def export_excel():
    """导出当前筛选条件的数据为 Excel"""
    # 复用 get_words 的筛选逻辑，但全量导出（不分页）
    conn = get_db()
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        # --- 与 /api/words 相同的筛选参数 ---
        freq_min = request.args.get('freq_min', type=int)
        freq_max = request.args.get('freq_max', type=int)
        years = request.args.get('years')
        year_count_min = request.args.get('year_count_min', type=int)
        year_count_max = request.args.get('year_count_max', type=int)
        first_year = request.args.get('first_year', type=int)
        exam_types = request.args.get('exam_types')
        search = request.args.get('search', type=str)
        sort_by = request.args.get('sort_by', 'freq')
        sort_order = request.args.get('sort_order', 'desc')

        conditions = []
        params = []

        if freq_min is not None:
            conditions.append("w.total_freq >= ?"); params.append(freq_min)
        if freq_max is not None:
            conditions.append("w.total_freq <= ?"); params.append(freq_max)
        if year_count_min is not None:
            conditions.append("w.year_count >= ?"); params.append(year_count_min)
        if year_count_max is not None:
            conditions.append("w.year_count <= ?"); params.append(year_count_max)
        if first_year is not None:
            conditions.append("w.first_year = ?"); params.append(first_year)
        if search and search.strip():
            conditions.append("w.word LIKE ?"); params.append(f"%{search.strip().lower()}%")
        if years and years.strip():
            year_list = [y.strip() for y in years.split(',') if y.strip()]
            if year_list:
                placeholders = ','.join(['?'] * len(year_list))
                conditions.append(f"w.id IN (SELECT DISTINCT word_id FROM occurrences WHERE year IN ({placeholders}))")
                params.extend(year_list)
        if exam_types and exam_types.strip():
            type_list = [t.strip() for t in exam_types.split(',') if t.strip()]
            if type_list:
                placeholders = ','.join(['?'] * len(type_list))
                conditions.append(f"w.id IN (SELECT DISTINCT word_id FROM occurrences WHERE exam_type IN ({placeholders}))")
                params.extend(type_list)

        sort_map = {'freq': 'w.total_freq', 'year_count': 'w.year_count',
                    'first_year': 'w.first_year', 'word': 'w.word'}
        sort_col = sort_map.get(sort_by, 'w.total_freq')
        order = 'DESC' if sort_order == 'desc' else 'ASC'

        where_clause = ' AND '.join(conditions) if conditions else '1=1'
        data_sql = f"""
            SELECT w.word, w.total_freq, w.year_count, w.first_year, w.last_year,
                   w.freq_by_year, w.freq_by_type
            FROM words w
            WHERE {where_clause}
            ORDER BY {sort_col} {order}
        """
        rows = conn.execute(data_sql, params).fetchall()

        # --- 获取所有年份和题型列表（用于列标题）---
        all_years = sorted([r[0] for r in
                           conn.execute("SELECT DISTINCT year FROM occurrences ORDER BY year")])
        all_types = sorted([r[0] for r in
                           conn.execute("SELECT DISTINCT exam_type FROM occurrences ORDER BY exam_type")])

        # --- 构建 Excel ---
        wb = Workbook()
        ws = wb.active
        ws.title = "词频统计"

        # 表头样式
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

        # 列：单词 | 总词频 | 年份数 | 首次年份 | 最后年份 | [每年词频] | [每题型词频]
        headers = ['单词', '总词频', '出现年份数', '首次年份', '最后年份']
        headers += [f'{y}年' for y in all_years]
        headers += all_types
        ws.append(headers)

        # 表头样式
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill

        # 数据行
        for row in rows:
            freq_by_year = json.loads(row['freq_by_year']) if row['freq_by_year'] else {}
            freq_by_type = json.loads(row['freq_by_type']) if row['freq_by_type'] else {}

            data_row = [
                row['word'],
                row['total_freq'],
                row['year_count'],
                row['first_year'],
                row['last_year']
            ]
            for y in all_years:
                data_row.append(freq_by_year.get(str(y), 0))
            for t in all_types:
                data_row.append(freq_by_type.get(t, 0))
            ws.append(data_row)

        # 自动筛选
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"

        # 列宽
        ws.column_dimensions['A'].width = 16
        for col_idx in range(2, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 10

        # 输出
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='英语真题词频统计.xlsx'
        )
    finally:
        conn.close()


# ============================================================
# 前端页面
# ============================================================
@app.route('/')
def index():
    """前端入口"""
    return render_template('index.html')


# ============================================================
# 启动
# ============================================================
if __name__ == '__main__':
    # Flask debug 模式会启动 reloader（父子两个进程）
    # 只在 reloader 子进程（实际运行 Flask 的进程）中打开浏览器
    if os.environ.get('WERKZEUG_RUN_MAIN') == 'true':
        threading.Timer(1.5, lambda: webbrowser.open('http://127.0.0.1:5000')).start()
    app.run(host='127.0.0.1', port=5000, debug=True)
